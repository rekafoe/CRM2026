# -*- coding: utf-8 -*-
"""Генерация Excel: себестоимость УФ-чернил на 1 м²."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "uv-ink-cost-per-m2.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Расчёт"

header_fill = PatternFill("solid", fgColor="D9E1F2")
bold = Font(bold=True)

ws["A1"] = "Себестоимость УФ-печати (чернила), руб / м²"
ws["A1"].font = Font(bold=True, size=12)

ws["A2"] = "Запас на промывки и потери, %"
ws["B2"] = 0
ws["C2"] = "добавляется к сумме чернил"

ws["A4"] = "Канал / чернила"
ws["B4"] = "Объём банки, мл"
ws["C4"] = "Цена банки, руб"
ws["D4"] = "Расход на 1 м², мл"
ws["E4"] = "Цена 1 мл, руб"
ws["F4"] = "Руб на 1 м² по каналу"

for c in range(1, 7):
    cell = ws.cell(row=4, column=c)
    cell.font = bold
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

default_channels = [
    "Cyan",
    "Magenta",
    "Yellow",
    "Black",
    "White",
    "Varnish / лак",
    "Glue / клей",
    "Другое",
]

start_row = 5
for i, name in enumerate(default_channels):
    r = start_row + i
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value="")
    ws.cell(row=r, column=3, value="")
    ws.cell(row=r, column=4, value="")
    ws.cell(row=r, column=5, value=f"=IF(B{r}=0,0,C{r}/B{r})")
    ws.cell(row=r, column=6, value=f"=IF(B{r}=0,0,D{r}*C{r}/B{r})")

last_data = start_row + len(default_channels) - 1

sum_row = last_data + 2
ws.cell(row=sum_row, column=1, value="Итого только чернила, руб/м²")
ws.cell(row=sum_row, column=1).font = bold
ws.cell(row=sum_row, column=6, value=f"=SUM(F{start_row}:F{last_data})")
ws.cell(row=sum_row, column=6).font = bold

adj_row = sum_row + 1
ws.cell(row=adj_row, column=1, value="Итого с учётом потерь, % (ячейка B2)")
ws.cell(row=adj_row, column=1).font = bold
ws.cell(row=adj_row, column=6, value=f"=F{sum_row}*(1+B2/100)")
ws.cell(row=adj_row, column=6).font = bold

mat_row = adj_row + 2
ws.cell(row=mat_row, column=1, value="Материал / подложка на 1 м², руб (вручную)")
ws.cell(row=mat_row, column=3, value=0)

tot_row = mat_row + 1
ws.cell(row=tot_row, column=1, value="ИТОГО чернила+потери+материал, руб/м²")
ws.cell(row=tot_row, column=1).font = bold
ws.cell(row=tot_row, column=6, value=f"=F{adj_row}+C{mat_row}")
ws.cell(row=tot_row, column=6).font = bold

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 22

for r in range(start_row, last_data + 1):
    for col in (2, 3, 4, 5, 6):
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="right")

ws2 = wb.create_sheet("Как пользоваться", 1)
help_lines = [
    "1. В столбце B укажите объём одной поставляемой банки (мл).",
    "2. В столбце C — цену этой банки в рублях (с НДС или без, как считаете для себя).",
    "3. В столбце D — ваш расход этого канала на 1 м² печати (мл). Берите из RIP, замеров или данных поставщика.",
    "4. Столбцы E и F считаются сами: цена за 1 мл и вклад канала в руб/м².",
    "5. Неиспользуемые строки оставьте пустыми (объём банки = 0 или не заполняйте расход).",
    "6. B2 — процент запаса на промывки и потери (например 5 означает +5% к сумме чернил).",
    "7. Материал на 1 м² — подложка, плёнка и т.п., если нужно в итог.",
    "",
    "Формула по каналу: (расход_мл_на_м² / объём_банки_мл) × цена_банки.",
]
for i, line in enumerate(help_lines, 1):
    ws2.cell(row=i, column=1, value=line)
ws2.column_dimensions["A"].width = 92

wb.save(OUT)
print("Saved:", OUT)
